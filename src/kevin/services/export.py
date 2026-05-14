from __future__ import annotations

import calendar
from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from kevin.repositories.asset import AssetRepository
from kevin.repositories.expense import ExpenseRepository
from kevin.repositories.income import IncomeRepository
from kevin.repositories.liability import LiabilityRepository


class ExportService:
    def __init__(
        self,
        expense_repo: ExpenseRepository,
        income_repo: IncomeRepository,
        asset_repo: AssetRepository,
        liability_repo: LiabilityRepository,
    ) -> None:
        self.expense_repo = expense_repo
        self.income_repo = income_repo
        self.asset_repo = asset_repo
        self.liability_repo = liability_repo

    def _in_range(
        self,
        year: int,
        month: int,
        start_year: int,
        start_month: int,
        end_year: int,
        end_month: int,
    ) -> bool:
        start = start_year * 12 + start_month
        end = end_year * 12 + end_month
        current = year * 12 + month
        return start <= current <= end

    def export_household(
        self,
        household_id: int,
        start_year: int,
        start_month: int,
        end_year: int,
        end_month: int,
        currency: str = "USD",
    ) -> bytes:
        wb = Workbook()

        # Collect all data for the range using list_by_year and filtering months
        years = range(start_year, end_year + 1)

        # --- Income sheet ---
        ws_income = wb.active
        ws_income.title = "Income"
        ws_income.append(["Year", "Month", "Title", f"Amount ({currency})"])

        income_items = []
        for year in years:
            for item in self.income_repo.list_by_year(household_id, year):
                if self._in_range(
                    item.year,
                    item.month,
                    start_year,
                    start_month,
                    end_year,
                    end_month,
                ):
                    income_items.append(item)

        income_items.sort(key=lambda x: (x.year, x.month))
        for item in income_items:
            ws_income.append(
                [item.year, calendar.month_name[item.month], item.title, item.amount]
            )

        # --- Expenses sheet ---
        ws_expenses = wb.create_sheet("Expenses")
        ws_expenses.append(["Year", "Month", "Title", f"Amount ({currency})"])

        expense_items = []
        for year in years:
            for item in self.expense_repo.list_by_year(household_id, year):
                if self._in_range(
                    item.year,
                    item.month,
                    start_year,
                    start_month,
                    end_year,
                    end_month,
                ):
                    expense_items.append(item)

        expense_items.sort(key=lambda x: (x.year, x.month))
        for item in expense_items:
            ws_expenses.append(
                [item.year, calendar.month_name[item.month], item.title, item.amount]
            )

        # --- Assets sheet ---
        ws_assets = wb.create_sheet("Assets")
        ws_assets.append(
            [
                "Year",
                "Month",
                "Title",
                "Ticker",
                "Quantity",
                f"Bought Price ({currency})",
                f"Current Price ({currency})",
                f"Value ({currency})",
            ]
        )

        asset_items = []
        for year in years:
            for item in self.asset_repo.list_by_year(household_id, year):
                if self._in_range(
                    item.year,
                    item.month,
                    start_year,
                    start_month,
                    end_year,
                    end_month,
                ):
                    asset_items.append(item)

        asset_items.sort(key=lambda x: (x.year, x.month))
        for item in asset_items:
            quantity = item.amount
            current_price = item.current_price
            value = (
                quantity * current_price
                if quantity is not None and current_price is not None
                else None
            )
            ws_assets.append(
                [
                    item.year,
                    calendar.month_name[item.month],
                    item.title,
                    item.ticker,
                    quantity,
                    item.bought_price,
                    current_price,
                    value,
                ]
            )

        # --- Liabilities sheet ---
        ws_liabilities = wb.create_sheet("Liabilities")
        ws_liabilities.append(["Year", "Month", "Title", f"Amount ({currency})"])

        liability_items = []
        for year in years:
            for item in self.liability_repo.list_by_year(household_id, year):
                if self._in_range(
                    item.year,
                    item.month,
                    start_year,
                    start_month,
                    end_year,
                    end_month,
                ):
                    liability_items.append(item)

        liability_items.sort(key=lambda x: (x.year, x.month))
        for item in liability_items:
            ws_liabilities.append(
                [item.year, calendar.month_name[item.month], item.title, item.amount]
            )

        # --- Summary sheet ---
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(
            [
                "Year",
                "Month",
                f"Total Income ({currency})",
                f"Total Expenses ({currency})",
                f"Net Savings ({currency})",
            ]
        )

        monthly_income: dict[tuple[int, int], Decimal] = defaultdict(Decimal)
        for item in income_items:
            monthly_income[(item.year, item.month)] += item.amount

        monthly_expense: dict[tuple[int, int], Decimal] = defaultdict(Decimal)
        for item in expense_items:
            monthly_expense[(item.year, item.month)] += item.amount

        all_months = sorted(set(monthly_income.keys()) | set(monthly_expense.keys()))
        for year, month in all_months:
            total_income = monthly_income[(year, month)]
            total_expenses = monthly_expense[(year, month)]
            net_savings = total_income - total_expenses
            ws_summary.append(
                [
                    year,
                    calendar.month_name[month],
                    total_income,
                    total_expenses,
                    net_savings,
                ]
            )

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
