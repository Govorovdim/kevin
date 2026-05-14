from __future__ import annotations

import calendar
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from kevin.models.asset import Asset
from kevin.models.expense import Expense
from kevin.models.income import Income
from kevin.models.liability import Liability
from kevin.repositories.asset import AssetRepository
from kevin.repositories.expense import ExpenseRepository
from kevin.repositories.income import IncomeRepository
from kevin.repositories.liability import LiabilityRepository

# Build month name -> number lookup: {"january": 1, "february": 2, ...}
MONTH_MAP = {name.lower(): num for num, name in enumerate(calendar.month_name) if num}


class ImportService:
    def __init__(
        self,
        expense_repo: ExpenseRepository,
        income_repo: IncomeRepository,
        asset_repo: AssetRepository,
        liability_repo: LiabilityRepository,
    ) -> None:
        self.expense_repo = expense_repo
        self.income_repo = income_repo
        self.asset_repo = asset_repo
        self.liability_repo = liability_repo

    def _parse_month(self, value: object) -> int:
        """Convert a month value (int, float, or name string) to an int 1-12."""
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            m = MONTH_MAP.get(value.strip().lower())
            if m is not None:
                return m
        raise ValueError(f"Invalid month: {value}")

    def import_household(self, household_id: int, file_bytes: bytes) -> dict:
        """Parse the xlsx and create records in a single transaction."""
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        session = self.income_repo.session  # all repos share the same session
        counts: dict[str, int] = {
            "income": 0,
            "expenses": 0,
            "assets": 0,
            "liabilities": 0,
        }

        # --- Income ---
        if "Income" in wb.sheetnames:
            ws = wb["Income"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                year, month_raw, title, amount = row[0], row[1], row[2], row[3]
                month = self._parse_month(month_raw)
                session.add(
                    Income(
                        household_id=household_id,
                        year=int(year),
                        month=month,
                        title=str(title),
                        amount=Decimal(str(amount)),
                    )
                )
                counts["income"] += 1

        # --- Expenses ---
        if "Expenses" in wb.sheetnames:
            ws = wb["Expenses"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                year, month_raw, title, amount = row[0], row[1], row[2], row[3]
                month = self._parse_month(month_raw)
                session.add(
                    Expense(
                        household_id=household_id,
                        year=int(year),
                        month=month,
                        title=str(title),
                        amount=Decimal(str(amount)),
                    )
                )
                counts["expenses"] += 1

        # --- Assets ---
        if "Assets" in wb.sheetnames:
            ws = wb["Assets"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                # Columns: Year, Month, Title, Ticker, Quantity, Bought Price, Current Price, Value
                year = row[0]
                month_raw = row[1]
                title = row[2]
                ticker = row[3] if len(row) > 3 else None
                quantity = row[4] if len(row) > 4 else None
                bought_price = row[5] if len(row) > 5 else None
                current_price = row[6] if len(row) > 6 else None
                # Value (col 7) is computed, skip it
                month = self._parse_month(month_raw)
                session.add(
                    Asset(
                        household_id=household_id,
                        year=int(year),
                        month=month,
                        title=str(title),
                        ticker=str(ticker) if ticker else None,
                        amount=Decimal(str(quantity)) if quantity is not None else None,
                        bought_price=Decimal(str(bought_price))
                        if bought_price is not None
                        else None,
                        current_price=Decimal(str(current_price))
                        if current_price is not None
                        else None,
                    )
                )
                counts["assets"] += 1

        # --- Liabilities ---
        if "Liabilities" in wb.sheetnames:
            ws = wb["Liabilities"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                year, month_raw, title, amount = row[0], row[1], row[2], row[3]
                month = self._parse_month(month_raw)
                session.add(
                    Liability(
                        household_id=household_id,
                        year=int(year),
                        month=month,
                        title=str(title),
                        amount=Decimal(str(amount)),
                    )
                )
                counts["liabilities"] += 1

        # Single commit for the entire import
        session.commit()
        wb.close()
        return counts
